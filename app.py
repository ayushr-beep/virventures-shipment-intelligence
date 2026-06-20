import streamlit as st
import pandas as pd
import numpy as np
import plotly.graph_objects as go
import io
from datetime import datetime

from modules import demand_engine, cost_model, decision_engine, excel_export, pptx_export, manifest_intelligence, lp_optimizer, stochastic_optimizer, data_continuity_checker, theme, data_persistence, box_packing_optimizer

st.set_page_config(page_title="Virventures | Master Shipment Intelligence", layout="wide", page_icon="📦")
theme.inject_theme(st)

REGIONS = ["East", "Central", "West"]
ORANGE = "#D2691E"
CHARCOAL = "#25272E"
SLATE = "#6B7280"
GREY = "#6B7280"
PANEL = "#F7F7F8"

# ---------------------------------------------------------------------------
# Left rail navigation (replaces top tabs)
# ---------------------------------------------------------------------------
active_page = theme.render_left_rail(st.sidebar)

# ---------------------------------------------------------------------------
# Data loading (runs every script execution regardless of active page, since
# every page after Setup needs this data — Streamlit reruns the whole script
# on every interaction, so this can't live "inside" just the Setup page body)
# ---------------------------------------------------------------------------
if "fee_schedule" not in st.session_state:
    st.session_state.fee_schedule = {
        k: dict(v) for k, v in cost_model.DEFAULT_PLACEMENT_FEE_SCHEDULE.items()
    }
if "use_sample" not in st.session_state:
    st.session_state.use_sample = True
if "window_days" not in st.session_state:
    st.session_state.window_days = 90
if "use_corrected" not in st.session_state:
    st.session_state.use_corrected = True


def _read(file_or_path, is_path=False):
    if is_path:
        return pd.read_excel(file_or_path)
    if file_or_path.name.endswith(".csv"):
        return pd.read_csv(file_or_path)
    return pd.read_excel(file_or_path)


@st.cache_data
def load_sample():
    sales = pd.read_excel("sample_data/sales_history_SAMPLE.xlsx")
    inventory = pd.read_excel("sample_data/inventory_by_region_SAMPLE.xlsx")
    invoices = pd.read_excel("sample_data/shipment_invoices_SAMPLE.xlsx")
    return sales, inventory, invoices


sales_df = inventory_df = invoices_df = None
data_ready = False
data_age_days = None
data_meta = None
use_sample = st.session_state.use_sample
window_days = st.session_state.window_days
use_corrected = st.session_state.use_corrected

if use_sample:
    sales_df, inventory_df, invoices_df = load_sample()
    data_ready = True
elif data_persistence.has_persisted_data():
    # Real data has been uploaded previously (e.g. this week's Monday refresh)
    # and persisted to disk -- load it automatically, no re-upload needed.
    try:
        sales_df, inventory_df, invoices_df, data_meta = data_persistence.load_persisted_data()
        data_ready = True
        data_age_days = data_persistence.get_data_age_days()
    except Exception as e:
        st.error(f"Couldn't read persisted data: {e}")
else:
    st.info(
        "No data uploaded yet. Go to **Data & Settings** in the left rail to upload your weekly "
        "Sales History, Inventory, and Invoices — they'll be saved automatically so you won't need "
        "to re-upload every session."
    )

# ---------------------------------------------------------------------------
# Main area
# ---------------------------------------------------------------------------
st.markdown(
    """
    <div style="background: var(--vv-near-black);
                border-radius: 18px; padding: 1.4rem 1.8rem; margin-bottom: 1.2rem;
                box-shadow: 0 8px 24px rgba(20,20,25,0.10);">
        <div style="font-size:0.76rem; font-weight:700; letter-spacing:0.08em; text-transform:uppercase; color:#D2691E;">
            VIRVENTURES &nbsp;·&nbsp; SHIPMENT INTELLIGENCE
        </div>
        <div style="font-family: Georgia, serif; font-size: 1.7rem; font-weight: 700; color:white; margin-top:0.15rem;">
            Master Shipment Intelligence Tool
        </div>
        <div style="font-size:0.88rem; color:#9CA0AC; margin-top:0.2rem;">
            Demand-weighted regional placement recommendations, with full cost transparency.
        </div>
    </div>
    """,
    unsafe_allow_html=True,
)

# Column mapping assumptions (sample + expected real schema use the same names;
# for a real upload with different headers, surface a mapping step).
expected_sales_cols = {"Order Date", "SKU", "ASIN", "Region", "Quantity"}
expected_inv_cols = {"SKU", "Region", "On-Hand Units"}
expected_inv_cols_fc = {"FC Code"}
expected_invoice_cols = {"Destination Region", "Total Units", "Total Weight (lb)", "Invoice Total ($)"}

if data_ready:
    missing_sales = expected_sales_cols - set(sales_df.columns)
    missing_inv = expected_inv_cols - set(inventory_df.columns)
    missing_invoice = expected_invoice_cols - set(invoices_df.columns)

    if missing_sales or missing_inv or missing_invoice:
        st.error(
            "Some expected columns weren't found. This version expects: \n\n"
            f"- Sales History: {sorted(expected_sales_cols)} (missing: {sorted(missing_sales) or 'none'})\n"
            f"- Inventory: {sorted(expected_inv_cols)} (missing: {sorted(missing_inv) or 'none'})\n"
            f"- Invoices: {sorted(expected_invoice_cols)} (missing: {sorted(missing_invoice) or 'none'})\n\n"
            "Rename columns to match, or let me know and I'll add a column-mapping step for your exact export format."
        )
        data_ready = False

# ---------------------------------------------------------------------------
# Compute engine outputs (recalculates live on any input change) — only when
# data is actually ready; the Setup page doesn't need any of this.
# ---------------------------------------------------------------------------
if data_ready:
    demand_profile = demand_engine.compute_demand_profile(sales_df, inventory_df, window_days=window_days)
    freight_rates = cost_model.derive_freight_rate_per_lb(invoices_df)
    as_of_date = demand_profile.attrs.get("as_of_date", "—")
    pct_basis_suffix = "corrected" if use_corrected else "raw"
    continuity_report = data_continuity_checker.run_full_continuity_report(sales_df)

if not data_ready and active_page != "setup":
    st.warning("No data loaded yet. Go to **Data & Settings** in the left rail to upload your files or use sample data.")
    st.stop()

# ---------------------------------------------------------------------------
# MANIFEST MODE — when active, every page below reads SCOPED versions of
# demand_profile / freight_rates / continuity_report instead of the full
# catalog. This is computed ONCE here, at the top level, so every page
# (Dashboard, Data Quality, Recommendation, Export — not just Shipment Plan)
# sees the same scoped data, rather than each page re-deriving its own
# partial version of "what SKUs are we talking about right now."
# ---------------------------------------------------------------------------
if "manifest_mode" not in st.session_state:
    st.session_state.manifest_mode = False
if "active_manifest" not in st.session_state:
    st.session_state.active_manifest = None  # dict: {parsed_df, raw_df, filename, summary_df, region_rows_df, unmatched_df, totals, packing_plan_df}

manifest_mode = st.session_state.manifest_mode and st.session_state.active_manifest is not None
manifest_scope_skus = []
demand_profile_scoped = demand_profile if data_ready else None
continuity_report_scoped = continuity_report if data_ready else None

if data_ready and manifest_mode:
    manifest_scope_skus = sorted(st.session_state.active_manifest["parsed_df"]["Merchant SKU"].astype(str).unique().tolist())
    scoped = demand_profile[demand_profile["SKU"].isin(manifest_scope_skus)]
    if len(scoped) > 0:
        demand_profile_scoped = scoped
        # Continuity check is recomputed on just the manifest's SKUs' sales
        # history, so its findings (gaps, anomalies) are specific to what
        # you're actually shipping right now, not diluted by your whole catalog.
        scoped_sales = sales_df[sales_df["SKU"].isin(manifest_scope_skus)]
        if len(scoped_sales) > 0:
            continuity_report_scoped = data_continuity_checker.run_full_continuity_report(scoped_sales)
    else:
        st.session_state.manifest_mode = False
        manifest_mode = False
        st.warning(
            "None of the active manifest's SKUs have sales history on record — turned Manifest Mode "
            "off automatically since there's nothing to scope to. Check the SKU spelling matches your "
            "sales data, or upload a different manifest."
        )

# ---------------------------------------------------------------------------
# Manifest Mode toggle bar — visible on every page, not just Shipment Plan.
# This is the one control that decides whether every tab below talks about
# your whole catalog or just the active manifest.
# ---------------------------------------------------------------------------
if data_ready and active_page != "setup":
    if st.session_state.active_manifest is not None:
        toggle_col1, toggle_col2 = st.columns([5, 2])
        with toggle_col1:
            manifest_name = st.session_state.active_manifest.get("filename", "uploaded manifest")
            toggle_label = (
                f"📦 Manifest Mode — scope everything to **{manifest_name}** ({len(manifest_scope_skus)} SKU(s))"
                if st.session_state.manifest_mode else
                f"📦 Manifest Mode — OFF (showing full catalog; manifest **{manifest_name}** is loaded and ready to activate)"
            )
            st.toggle(toggle_label, key="manifest_mode")
        with toggle_col2:
            if st.button("Clear manifest", key="_clear_manifest_top"):
                st.session_state.active_manifest = None
                # Can't assign to st.session_state.manifest_mode here -- the
                # toggle widget bound to that key has already instantiated
                # earlier in this same script run, and Streamlit forbids
                # writing to a widget-bound key after the widget exists.
                # Deleting the key instead lets it fall back to the default
                # (False) cleanly on the rerun below.
                del st.session_state["manifest_mode"]
                st.rerun()
    else:
        st.caption("📦 No manifest loaded — upload one on the **Shipment Plan** page to enable Manifest Mode.")

# ============================ DASHBOARD PAGE ===============================
if active_page == "dashboard":
    dash_df = demand_profile_scoped
    dash_continuity = continuity_report_scoped

    if manifest_mode:
        st.success(
            f"📦 **Manifest Mode active** — every chart and number below reflects only the "
            f"{len(manifest_scope_skus)} SKU(s) on **{st.session_state.active_manifest.get('filename', 'your manifest')}**, "
            f"not your full catalog."
        )

    if dash_continuity["total_span_months"] >= 12 and not dash_continuity["is_ready_for_seasonality_modeling"]:
        reasons_str = "; ".join(dash_continuity.get("not_ready_reasons", []))
        st.warning(
            f"⚠️ Data quality check found issues worth reviewing before trusting long-range trends: "
            f"{reasons_str}. See **Data Quality** in the left rail for details before relying on seasonal patterns."
        )

    col1, col2, col3, col4 = st.columns(4)
    total_units_sold = sum(dash_df[f"{r}_units_sold"].sum() for r in REGIONS)
    total_on_hand = sum(dash_df[f"{r}_on_hand"].sum() for r in REGIONS)
    east_share_current = dash_df["East_on_hand"].sum() / total_on_hand if total_on_hand else 0
    n_skus_shown = len(dash_df)

    col1.metric("SKUs in View", f"{n_skus_shown}")
    col2.metric("Units Sold (window)", f"{int(total_units_sold):,}")
    col3.metric("Current On-Hand", f"{int(total_on_hand):,}")
    col4.metric("Current East Inventory Share", f"{east_share_current*100:.0f}%",
                help="Your stated status quo: majority of inventory sits in East today.")

    chart_col1, chart_col2 = st.columns([3, 2])

    with chart_col1:
        st.markdown("#### Regional Demand vs. Current Inventory Allocation")
        st.caption(
            f"Demand basis: {'sell-through-corrected (velocity-weighted)' if use_corrected else 'raw unit volume'}. "
            "If a region's demand bar is taller than its inventory bar, that region is likely under-stocked."
        )

        agg_demand = {r: dash_df[f"{r}_demand_pct_{pct_basis_suffix}"].mean() for r in REGIONS}
        agg_inventory_share = {r: dash_df[f"{r}_on_hand"].sum() / total_on_hand if total_on_hand else 0 for r in REGIONS}

        fig = go.Figure()
        fig.add_trace(go.Bar(
            name="Demand Share", x=REGIONS, y=[agg_demand[r]*100 for r in REGIONS],
            marker_color=ORANGE, text=[f"{agg_demand[r]*100:.0f}%" for r in REGIONS], textposition="outside"
        ))
        fig.add_trace(go.Bar(
            name="Current Inventory Share", x=REGIONS, y=[agg_inventory_share[r]*100 for r in REGIONS],
            marker_color=GREY, text=[f"{agg_inventory_share[r]*100:.0f}%" for r in REGIONS], textposition="outside"
        ))
        fig.update_layout(
            barmode="group", height=380, plot_bgcolor="white", paper_bgcolor="white",
            font=dict(family="Georgia, serif", color=CHARCOAL),
            yaxis_title="% share", legend=dict(orientation="h", yanchor="bottom", y=1.02),
            margin=dict(t=40, b=20, l=40, r=20),
        )
        st.plotly_chart(fig, width='stretch')

    with chart_col2:
        st.markdown("#### Current Inventory Mix")
        st.caption("Where your on-hand units actually sit today, across the SKUs in view.")
        inv_pie = go.Figure(data=[go.Pie(
            labels=REGIONS,
            values=[dash_df[f"{r}_on_hand"].sum() for r in REGIONS],
            marker=dict(colors=[ORANGE, "#9CA3AF", CHARCOAL]),
            hole=0.45,
            textinfo="label+percent",
        )])
        inv_pie.update_layout(
            height=380, showlegend=False, font=dict(family="Georgia, serif", color=CHARCOAL),
            margin=dict(t=20, b=20, l=20, r=20),
        )
        st.plotly_chart(inv_pie, width='stretch')

    st.markdown("#### Demand Split Matrix — Every SKU × Every Region")
    if manifest_mode:
        st.caption(f"📦 Scoped to the {n_skus_shown} SKU(s) on your active manifest.")
    st.caption(
        "Each row is a SKU, each column a region, each cell its demand share. Darker orange = more "
        "demand concentrated there. Scan down a column to see which SKUs lean toward that region; "
        "scan across a row to see how concentrated or spread out one SKU's demand is."
    )

    matrix_cols = [f"{r}_demand_pct_{pct_basis_suffix}" for r in REGIONS]
    matrix_df = dash_df[["SKU"] + matrix_cols].copy()
    matrix_df.columns = ["SKU"] + REGIONS
    for r in REGIONS:
        matrix_df[r] = (matrix_df[r] * 100).round(1)
    matrix_df = matrix_df.set_index("SKU")

    fig_matrix = go.Figure(data=go.Heatmap(
        z=matrix_df.values,
        x=REGIONS,
        y=matrix_df.index.tolist(),
        colorscale=[[0, "#FFFFFF"], [0.5, "#F0997B"], [1, "#993C1D"]],
        text=[[f"{v:.0f}%" for v in row] for row in matrix_df.values],
        texttemplate="%{text}",
        textfont=dict(size=13, family="Georgia, serif", color=CHARCOAL),
        colorbar=dict(title="Demand %", ticksuffix="%"),
        zmin=0, zmax=max(60, matrix_df.values.max()),
        hovertemplate="SKU: %{y}<br>Region: %{x}<br>Demand: %{z:.1f}%<extra></extra>",
    ))
    fig_matrix.update_layout(
        height=max(220, 60 + 38 * len(matrix_df)),
        plot_bgcolor="white", paper_bgcolor="white",
        font=dict(family="Georgia, serif", color=CHARCOAL),
        margin=dict(t=20, b=20, l=20, r=20),
        xaxis=dict(side="top"),
        yaxis=dict(autorange="reversed"),
    )
    st.plotly_chart(fig_matrix, width='stretch')

    with st.expander("View as plain table instead"):
        plain_table = matrix_df.reset_index()
        plain_table.columns = ["SKU"] + [f"{r} Demand %" for r in REGIONS]
        st.dataframe(plain_table, width='stretch', hide_index=True)

    st.markdown("#### Freight Rate by Region (back-calculated from your invoices)")
    freight_chart_col, freight_table_col = st.columns([2, 3])
    with freight_chart_col:
        freight_fig = go.Figure(data=[go.Bar(
            x=freight_rates["Region"], y=freight_rates["rate_per_lb"],
            marker_color=[ORANGE if r == freight_rates.loc[freight_rates["rate_per_lb"].idxmin(), "Region"] else GREY for r in freight_rates["Region"]],
            text=[f"${v:.2f}" for v in freight_rates["rate_per_lb"]], textposition="outside",
        )])
        freight_fig.update_layout(
            height=300, plot_bgcolor="white", paper_bgcolor="white",
            font=dict(family="Georgia, serif", color=CHARCOAL),
            yaxis_title="$/lb", margin=dict(t=20, b=20, l=40, r=20), showlegend=False,
        )
        st.plotly_chart(freight_fig, width='stretch')
    with freight_table_col:
        rate_display = freight_rates[["Region", "rate_per_lb", "rate_per_unit", "n_invoices"]].copy()
        rate_display.columns = ["Region", "$/lb", "$/unit", "Invoices Used"]
        rate_display["$/lb"] = rate_display["$/lb"].round(2)
        rate_display["$/unit"] = rate_display["$/unit"].round(2)
        st.dataframe(rate_display, width='stretch', hide_index=True)
        if (freight_rates["n_invoices"] == 0).any():
            st.caption("⚠️ Some regions have 0 invoices on record — their rate falls back to the average of "
                       "other regions. Treat that number as a rough placeholder, not a derived rate.")

# ============================ DATA QUALITY PAGE =============================
if active_page == "quality":
    if manifest_mode:
        st.success(
            f"📦 **Manifest Mode active** — continuity checks below run only against the "
            f"{len(manifest_scope_skus)} SKU(s) on your active manifest."
        )
    st.markdown("#### Data Continuity Check")
    st.caption(
        "Runs automatically on every upload, BEFORE any long-range trend or seasonality conclusions "
        "are drawn from your data. This flags problems for you to judge — it doesn't auto-fix anything, "
        "because deciding whether a gap is a real business event or a data export issue needs a human "
        "who knows the history."
    )

    cq1, cq2, cq3, cq4 = st.columns(4)
    cq1.metric("History Span", f"{continuity_report_scoped['total_span_months']} months")
    cq2.metric("Timeline Gaps", continuity_report_scoped["n_timeline_gaps"])
    cq3.metric("SKUs Needing Review", continuity_report_scoped["n_skus_needing_review"])
    cq4.metric("Anomalous Periods", continuity_report_scoped["n_anomalous_periods"])

    if continuity_report_scoped["is_ready_for_seasonality_modeling"]:
        st.success(
            "✅ This data looks continuous enough for seasonality/trend modeling: 24+ months of history, "
            "no timeline gaps, no SKU discontinuation patterns detected."
        )
    else:
        st.warning(
            "This data has issues worth reviewing before trusting long-range seasonal patterns built on it. "
            "Specifics below — none of these block you from using the tool normally (the Recommendation "
            "and Shipment Plan tabs work fine regardless), this only matters for trusting multi-year trend "
            "claims."
        )
        for reason in continuity_report_scoped.get("not_ready_reasons", []):
            st.markdown(f"- {reason}")

    with st.expander("Timeline gaps — months with suspiciously low or zero order volume across ALL SKUs"):
        st.caption(
            "It's extremely unlikely every SKU genuinely has zero sales in the same month — this usually "
            "means a data export gap, an account migration, or a reporting format change, not a real "
            "demand collapse."
        )
        timeline_flagged = continuity_report_scoped["timeline_detail"][continuity_report_scoped["timeline_detail"]["flag"]]
        if len(timeline_flagged) > 0:
            display_timeline = timeline_flagged[["YearMonth", "order_count", "flag_reason"]].copy()
            display_timeline["YearMonth"] = display_timeline["YearMonth"].astype(str)
            st.dataframe(display_timeline, width='stretch', hide_index=True)
        else:
            st.caption("None found.")

    with st.expander("SKU continuity — products that went quiet for a while, then came back"):
        st.caption(
            "A SKU with a multi-month gap sandwiched between active periods may have been discontinued "
            "and relaunched, had a listing/variation change, or the SKU code got reused for a different "
            "product. Verify before treating its pre- and post-gap history as one continuous pattern."
        )
        sku_gaps_flagged = continuity_report_scoped["sku_gap_detail"][continuity_report_scoped["sku_gap_detail"]["needs_review"]]
        if len(sku_gaps_flagged) > 0:
            st.dataframe(sku_gaps_flagged, width='stretch', hide_index=True)
        else:
            st.caption("None found.")
        with st.popover("View all SKUs (including ones with no issues)"):
            st.dataframe(continuity_report_scoped["sku_gap_detail"], width='stretch', hide_index=True)

    with st.expander("Anomalous periods — months far outside a SKU's normal range"):
        st.caption(
            "Statistical outliers relative to each SKU's own history (e.g. a demand shock, viral spike, "
            "or stockout-driven collapse). These would distort a seasonal average if blended in without "
            "knowing they're there — doesn't mean exclude them, just know they exist."
        )
        if len(continuity_report_scoped["anomaly_detail"]) > 0:
            st.dataframe(continuity_report_scoped["anomaly_detail"], width='stretch', hide_index=True)
        else:
            st.caption("None found.")

# ============================ RECOMMENDATION PAGE ===========================
if active_page == "recommend":
    st.markdown("#### Get a Recommendation for a Specific Shipment")

    if manifest_mode:
        st.success(
            f"📦 **Manifest Mode active** — SKU list below shows only the {len(manifest_scope_skus)} "
            f"SKU(s) on **{st.session_state.active_manifest.get('filename', 'your manifest')}**."
        )

    sku_options = demand_profile_scoped["SKU"].tolist()
    selected_sku = st.selectbox("Select SKU", sku_options)
    total_units_input = st.number_input("Total units in this shipment", min_value=1, value=1000, step=50)

    c1, c2 = st.columns(2)
    size_tier = c1.selectbox("Size tier", list(st.session_state.fee_schedule.keys()), index=1)
    avg_unit_weight = c2.number_input("Avg unit weight (lb)", min_value=0.01, value=1.1, step=0.1)

    sku_row = demand_profile_scoped[demand_profile_scoped["SKU"] == selected_sku].iloc[0]
    if use_corrected:
        demand_pct = demand_engine.get_sku_demand_summary(sku_row)
    else:
        demand_pct = demand_engine.get_sku_demand_summary_raw(sku_row)

    engine_mode = st.radio(
        "Recommendation method",
        [
            "Quick heuristic (proportional split)",
            "Optimized (linear programming — provably minimal cost)",
            "Stochastic (accounts for demand uncertainty across history)",
        ],
        horizontal=True,
        help="The heuristic splits units proportionally to demand, then checks the cost. The LP optimizer "
             "searches every possible split to find the mathematically cheapest option for ONE assumed "
             "demand%. The stochastic optimizer builds several scenarios from your actual sales history "
             "(demand% as it looked in different historical months) and finds a split that performs well "
             "across all of them — either cheapest on average, or safest against the worst month."
    )

    if engine_mode.startswith("Stochastic"):
        st.markdown("##### Build Demand Scenarios from History")
        sc1, sc2 = st.columns(2)
        n_scenarios_requested = sc1.slider(
            "Number of historical scenarios to build", min_value=2, max_value=12, value=6,
            help="Each scenario is a separate trailing-90-day window taken at a different point in your "
                 "sales history — e.g. roughly one per month if you have 6+ months of data."
        )
        stoch_mode = sc2.radio(
            "Optimize for", ["Expected cost (cheapest on average)", "Worst case (safest against a bad month)"],
            help="Expected mode minimizes average cost across all scenarios. Worst-case mode minimizes "
                 "the maximum cost/shortfall across any single scenario — costs a bit more on average, "
                 "but bounds your downside if demand looks like your worst historical month."
        )

        scenarios, n_actual = stochastic_optimizer.build_demand_scenarios(
            sales_df, inventory_df, selected_sku, window_days=90, n_scenarios=n_scenarios_requested
        )

        if n_actual == 0:
            st.error(f"No sales history found for {selected_sku} — can't build scenarios.")
            st.stop()
        if n_actual < n_scenarios_requested:
            st.warning(
                f"Only {n_actual} distinct historical windows available (requested {n_scenarios_requested}) "
                f"— there isn't enough sales history to build more non-overlapping scenarios. Results are "
                f"still valid, just based on fewer data points than requested."
            )

        with st.expander(f"View the {n_actual} scenarios used"):
            scenario_display = pd.DataFrame([
                {"As-Of Date": str(s["window_end"])[:10],
                 "East %": round(s["East"]*100, 1), "Central %": round(s["Central"]*100, 1),
                 "West %": round(s["West"]*100, 1)}
                for s in scenarios
            ])
            st.dataframe(scenario_display, width='stretch', hide_index=True)

        mode_key = "worst_case" if stoch_mode.startswith("Worst") else "expected"
        stoch_result = stochastic_optimizer.solve_stochastic_split(
            total_units_input, scenarios, freight_rates, size_tier, avg_unit_weight,
            fee_schedule=st.session_state.fee_schedule, mode=mode_key
        )

        if not stoch_result["is_optimal"]:
            st.error(f"Stochastic solver failed: {stoch_result.get('error', 'unknown error')}")
            st.stop()

        st.markdown("##### Recommended Split (robust to demand uncertainty)")
        rcols = st.columns(3)
        for i, region in enumerate(REGIONS):
            units = stoch_result["units"][region]
            is_top = units == max(stoch_result["units"].values())
            with rcols[i]:
                theme.render_region_tile(st, region, units, "", is_top)

        st.markdown("##### Performance Across Historical Scenarios")
        cc1, cc2, cc3 = st.columns(3)
        cc1.metric("Fixed Cost (freight + placement)", f"${stoch_result['fixed_cost']:,.2f}")
        cc2.metric("Demand Served — Worst Scenario", f"{stoch_result['min_demand_served_pct']:.1f}%")
        cc3.metric("Demand Served — Best Scenario", f"{stoch_result['max_demand_served_pct']:.1f}%")

        st.caption(
            f"Worst-case unit shortfall across all {n_actual} scenarios: "
            f"{stoch_result['worst_case_unit_shortfall']:.0f} units. This is the number 'worst case' "
            f"mode is directly minimizing — switch modes above to compare."
        )

        per_scenario_df = pd.DataFrame([
            {
                "Scenario (as of)": s["window_end"][:10],
                "East %": round(s["demand_pct"]["East"]*100, 1),
                "Central %": round(s["demand_pct"]["Central"]*100, 1),
                "West %": round(s["demand_pct"]["West"]*100, 1),
                "Unit Shortfall": s["unit_shortfall_total"],
                "Demand Served %": s["demand_served_pct"],
            }
            for s in stoch_result["per_scenario"]
        ])
        st.dataframe(per_scenario_df, width='stretch', hide_index=True)

        st.session_state["_last_recommendation"] = {
            "sku": selected_sku, "region_split": stoch_result["units"], "demand_pct": demand_pct,
            "recommended_cost": stoch_result["fixed_cost"],
            "cheapest_region": min(freight_rates.set_index("Region")["rate_per_lb"].to_dict(),
                                    key=lambda r: freight_rates.set_index("Region")["rate_per_lb"].to_dict()[r]),
            "cheapest_cost": stoch_result["fixed_cost"],
            "cost_delta": 0,
            "coverage_gap_pct": 100 - stoch_result["min_demand_served_pct"],
            "rationale": (
                f"Stochastic split ({mode_key} mode) across {n_actual} historical demand scenarios: "
                f"${stoch_result['fixed_cost']:,.2f} fixed cost, demand served ranges from "
                f"{stoch_result['min_demand_served_pct']:.1f}% (worst scenario) to "
                f"{stoch_result['max_demand_served_pct']:.1f}% (best scenario) depending on which "
                f"historical demand pattern actually recurs."
            ),
        }

    elif engine_mode.startswith("Optimized"):
        min_coverage_pct = st.slider(
            "Minimum demand coverage required (%)", min_value=0, max_value=100, value=100, step=5,
            help="0% = pure cost minimization, may ship everything to one region regardless of demand. "
                 "100% = every region with demand must receive a proportional share. Move this slider to "
                 "see the real cost-vs-coverage tradeoff, not just two fixed points."
        ) / 100.0

        lp_result = lp_optimizer.optimize_split(
            total_units_input, demand_pct, freight_rates, size_tier, avg_unit_weight,
            fee_schedule=st.session_state.fee_schedule, min_coverage_pct=min_coverage_pct
        )

        if not lp_result["is_optimal"]:
            st.error(
                f"The optimizer couldn't find a feasible solution at {min_coverage_pct*100:.0f}% coverage "
                f"for this shipment size. {lp_result.get('error', '')} Try lowering the coverage requirement."
            )
            st.stop()

        display_units = lp_result["units"]
        display_cost = lp_result["cost"]["total_cost"]
        display_n_locations = lp_result["cost"]["n_locations"]

        st.markdown("##### Optimized Split")
        rcols = st.columns(3)
        for i, region in enumerate(REGIONS):
            units = display_units[region]
            pct = demand_pct.get(region, 0) * 100
            is_top = units == max(display_units.values())
            with rcols[i]:
                theme.render_region_tile(st, region, units, f"{pct:.0f}% of demand", is_top)

        st.caption(f"Uses {display_n_locations} location(s) — solver status: {lp_result['solver_status']}")

        st.markdown("##### Cost-vs-Coverage Tradeoff Curve")
        st.caption(
            "This is the actual Pareto frontier — every point is a provably optimal split AT that coverage "
            "level. Moving the slider above traces this exact curve, so you can see precisely what each "
            "percentage point of demand coverage costs."
        )
        curve_df = lp_optimizer.compare_coverage_tradeoff_curve(
            total_units_input, demand_pct, freight_rates, size_tier, avg_unit_weight,
            fee_schedule=st.session_state.fee_schedule
        )
        fig_curve = go.Figure()
        fig_curve.add_trace(go.Scatter(
            x=curve_df["Min Coverage Required (%)"], y=curve_df["Total Cost ($)"],
            mode="lines+markers", line=dict(color=ORANGE, width=2.5), marker=dict(size=8),
        ))
        fig_curve.add_trace(go.Scatter(
            x=[min_coverage_pct * 100], y=[display_cost],
            mode="markers", marker=dict(size=14, color=CHARCOAL, symbol="diamond"),
            name="Your current selection", showlegend=False,
        ))
        fig_curve.update_layout(
            height=320, plot_bgcolor="white", paper_bgcolor="white",
            font=dict(family="Georgia, serif", color=CHARCOAL),
            xaxis_title="Minimum demand coverage required (%)", yaxis_title="Total landed cost ($)",
            margin=dict(t=20, b=40, l=60, r=20),
        )
        st.plotly_chart(fig_curve, width='stretch')

        cheapest_region_lp = min(
            freight_rates.set_index("Region")["rate_per_lb"].to_dict(),
            key=lambda r: freight_rates.set_index("Region")["rate_per_lb"].to_dict()[r]
        )
        cheapest_cost_lp = curve_df.iloc[0]["Total Cost ($)"]
        cost_delta_lp = display_cost - cheapest_cost_lp

        # Real demand-coverage gap (percentage points of demand NOT served by
        # shipping only to the cheapest region) -- same definition used by the
        # heuristic path, NOT a cost ratio. Mixing these up would put a
        # misleading number in front of the person reading the export.
        demand_coverage_lp = sum(demand_pct.get(r, 0) for r, u in display_units.items() if u > 0)
        cheapest_coverage_lp = demand_pct.get(cheapest_region_lp, 0)
        coverage_gap_pct_lp = (demand_coverage_lp - cheapest_coverage_lp) * 100

        st.session_state["_last_recommendation"] = {
            "sku": selected_sku, "region_split": display_units, "demand_pct": demand_pct,
            "recommended_cost": display_cost,
            "cheapest_region": cheapest_region_lp,
            "cheapest_cost": cheapest_cost_lp,
            "cost_delta": cost_delta_lp,
            "coverage_gap_pct": coverage_gap_pct_lp,
            "rationale": (
                f"LP-optimized split at {min_coverage_pct*100:.0f}% minimum demand coverage: "
                f"${display_cost:,.2f} total landed cost using {display_n_locations} location(s). "
                f"At 0% coverage (pure cost minimization), the cheapest possible cost is "
                f"${cheapest_cost_lp:,.2f} -- the ${cost_delta_lp:,.2f} difference is the price of "
                f"covering demand outside {cheapest_region_lp}."
            ),
        }

    else:
        result = decision_engine.recommend_split(
            total_units_input, demand_pct, freight_rates, size_tier, avg_unit_weight,
            fee_schedule=st.session_state.fee_schedule
        )

        st.markdown("##### Recommended Split")
        rcols = st.columns(3)
        for i, region in enumerate(REGIONS):
            units = result["demand_optimal"]["units"][region]
            pct = demand_pct.get(region, 0) * 100
            is_top = units == max(result["demand_optimal"]["units"].values())
            with rcols[i]:
                theme.render_region_tile(st, region, units, f"{pct:.0f}% of demand", is_top)

        st.markdown("##### Cost Tradeoff")
        cc1, cc2, cc3 = st.columns(3)
        cc1.metric("Recommended Split Cost", f"${result['demand_optimal']['cost']['total_cost']:,.2f}")
        cc2.metric(f"Cheapest Option ({result['cheapest']['region']} only)",
                   f"${result['cheapest']['cost']['total_cost']:,.2f}")
        delta = result["cost_delta_vs_cheapest"]
        cc3.metric("Cost Delta", f"${abs(delta):,.2f} {'more' if delta > 0 else 'less' if delta < 0 else 'same'}",
                   delta=f"{delta:,.2f}", delta_color="inverse")

        st.info(result["rationale"])

        st.session_state["_last_recommendation"] = {
            "sku": selected_sku, "region_split": result["demand_optimal"]["units"], "demand_pct": demand_pct,
            "recommended_cost": result["demand_optimal"]["cost"]["total_cost"],
            "cheapest_region": result["cheapest"]["region"],
            "cheapest_cost": result["cheapest"]["cost"]["total_cost"],
            "cost_delta": result["cost_delta_vs_cheapest"],
            "coverage_gap_pct": result["coverage_gap_pct"] * 100,
            "rationale": result["rationale"],
        }

# ============================ MANIFEST / SHIPMENT PLAN PAGE ==================
if active_page == "manifest":
    st.markdown("#### Upload a Send to Amazon Manifest")
    st.caption(
        "Upload the actual manifest you'd send to Seller Central (the 'Create workflow' template, "
        "with Merchant SKU + Quantity columns). The tool reads every SKU on it and runs the same "
        "demand + cost engine used in the Recommendation tab — for the whole shipment at once. "
        "Once processed, flip on **Manifest Mode** (top of any page) to scope the entire app to it."
    )

    manifest_file = st.file_uploader("Send to Amazon manifest (.xlsx)", type=["xlsx"], key="manifest_upload")

    m1, m2 = st.columns(2)
    manifest_size_tier = m1.selectbox("Size tier (applies to all SKUs in this manifest)",
                                       list(st.session_state.fee_schedule.keys()), index=1, key="manifest_tier")
    manifest_weight = m2.number_input("Avg unit weight (lb, applies to all SKUs)", min_value=0.01,
                                       value=1.1, step=0.1, key="manifest_weight")

    if manifest_file is not None:
        try:
            manifest_bytes = manifest_file.getvalue()
            parsed_manifest = manifest_intelligence.parse_manifest(io.BytesIO(manifest_bytes))
            raw_manifest_for_boxes = manifest_intelligence.parse_manifest_full(io.BytesIO(manifest_bytes))
        except ValueError as e:
            st.error(str(e))
            parsed_manifest = None
            raw_manifest_for_boxes = None

        if parsed_manifest is not None:
            st.success(f"Read {len(parsed_manifest)} SKU lines from the manifest.")

            summary_df, region_rows_df, unmatched_df, totals = manifest_intelligence.build_shipment_plan(
                parsed_manifest, demand_profile, freight_rates, manifest_size_tier, manifest_weight,
                fee_schedule=st.session_state.fee_schedule
            )

            box_specs_by_sku = box_packing_optimizer.extract_box_specs(raw_manifest_for_boxes) if raw_manifest_for_boxes is not None else {}
            packing_plan_df = box_packing_optimizer.build_full_packing_plan(
                region_rows_df, box_specs_by_sku, manifest_size_tier, fee_schedule=st.session_state.fee_schedule
            ) if len(region_rows_df) > 0 else pd.DataFrame()

            # Persist everything into session_state.active_manifest -- this is
            # what Manifest Mode reads from on every other page, and what
            # survives navigating away and back without re-uploading.
            st.session_state.active_manifest = {
                "filename": getattr(manifest_file, "name", "uploaded manifest"),
                "parsed_df": parsed_manifest,
                "raw_df": raw_manifest_for_boxes,
                "summary_df": summary_df,
                "region_rows_df": region_rows_df,
                "unmatched_df": unmatched_df,
                "totals": totals,
                "packing_plan_df": packing_plan_df,
                "box_specs_by_sku": box_specs_by_sku,
                "size_tier": manifest_size_tier,
                "avg_unit_weight": manifest_weight,
            }

            if not st.session_state.manifest_mode and len(summary_df) > 0:
                st.info(
                    "💡 Manifest processed. Flip on **Manifest Mode** at the top of any page to scope "
                    "the whole app — Dashboard, Recommendation, Export, everything — to just this shipment."
                )

            if totals["unmatched_skus"] > 0:
                st.warning(
                    f"⚠️ {totals['unmatched_skus']} of {len(parsed_manifest)} SKUs on this manifest have "
                    f"no sales history in your uploaded data — there's no demand signal for them, so they're "
                    f"excluded from the recommendation below. Review them in the 'Unmatched SKUs' section and "
                    f"either ship them by your existing process, or check the SKU spelling matches your sales data."
                )

            if len(summary_df) > 0:
                st.markdown("##### Shipment Totals")
                t1, t2, t3, t4 = st.columns(4)
                t1.metric("SKUs Matched", f"{totals['total_skus']}")
                t2.metric("Total Units", f"{totals['total_units']:,}")
                t3.metric("Recommended Cost", f"${totals['total_recommended_cost']:,.2f}")
                delta = totals["total_cost_delta"]
                t4.metric("Cost Delta vs. Cheapest-Only",
                          f"${abs(delta):,.2f} {'more' if delta > 0 else 'less' if delta < 0 else 'same'}",
                          delta=f"{delta:,.2f}", delta_color="inverse")

                viz_col1, viz_col2 = st.columns(2)
                with viz_col1:
                    st.markdown("###### Units by Region")
                    region_totals = region_rows_df.groupby("Region")["Quantity"].sum().reindex(REGIONS).fillna(0)
                    region_bar = go.Figure(data=[go.Bar(
                        x=REGIONS, y=region_totals.values,
                        marker_color=[ORANGE if v == region_totals.max() else GREY for v in region_totals.values],
                        text=[f"{int(v):,}" for v in region_totals.values], textposition="outside",
                    )])
                    region_bar.update_layout(
                        height=300, plot_bgcolor="white", paper_bgcolor="white",
                        font=dict(family="Georgia, serif", color=CHARCOAL),
                        margin=dict(t=20, b=20, l=40, r=20), showlegend=False,
                    )
                    st.plotly_chart(region_bar, width='stretch')
                with viz_col2:
                    st.markdown("###### Cost Split by SKU")
                    cost_pie = go.Figure(data=[go.Pie(
                        labels=summary_df["Merchant SKU"], values=summary_df["Recommended Cost ($)"],
                        marker=dict(colors=[ORANGE, "#9CA3AF", CHARCOAL, "#E8915A", "#5B6168"][:len(summary_df)]),
                        hole=0.45, textinfo="label+percent",
                    )])
                    cost_pie.update_layout(
                        height=300, showlegend=False, font=dict(family="Georgia, serif", color=CHARCOAL),
                        margin=dict(t=20, b=20, l=20, r=20),
                    )
                    st.plotly_chart(cost_pie, width='stretch')

                view_mode = st.radio(
                    "View", ["Summary (one row per SKU)", "Region rows (one row per SKU per region)"],
                    horizontal=True
                )

                if view_mode.startswith("Summary"):
                    st.dataframe(summary_df, width='stretch', hide_index=True)
                else:
                    st.dataframe(region_rows_df, width='stretch', hide_index=True)

                st.markdown("##### Per-Region Send-to-Amazon Files — Exact Manifest Format")
                st.caption(
                    "Real .xlsx files matching the Send-to-Amazon template exactly (same sheet name, same "
                    "header layout) — open and re-upload directly, no reformatting needed."
                )
                region_cols = st.columns(3)
                for i, region in enumerate(REGIONS):
                    region_export = manifest_intelligence.build_region_manifest_export(region_rows_df, region)
                    if len(region_export) == 0:
                        region_cols[i].caption(f"No units allocated to {region} in this shipment.")
                        continue
                    region_xlsx_bytes = manifest_intelligence.build_exact_format_manifest_xlsx(region_export)
                    region_cols[i].download_button(
                        f"⬇️ {region} ({region_export['Quantity'].sum()} units)",
                        data=region_xlsx_bytes,
                        file_name=f"manifest_{region.lower()}_{datetime.now().strftime('%Y%m%d')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key=f"manifest_dl_{region}",
                    )

                st.markdown("---")
                st.markdown("##### Box-Level Packing — How Each Region's Units Become Boxes")
                st.caption(
                    "Each region's unit allocation, broken into actual boxes using the units-per-box and "
                    "box dimensions/weight from your manifest. SKUs with no box spec on file fall back to "
                    "an editable default rather than guessing silently."
                )

                with st.expander("Default box spec (used only for SKUs with no box info on the manifest)", expanded=False):
                    bc1, bc2, bc3, bc4 = st.columns(4)
                    box_packing_optimizer.DEFAULT_BOX_SPEC["units_per_box"] = bc1.number_input(
                        "Units per box", value=box_packing_optimizer.DEFAULT_BOX_SPEC["units_per_box"], min_value=1, key="default_units_per_box"
                    )
                    box_packing_optimizer.DEFAULT_BOX_SPEC["box_weight_lb"] = bc2.number_input(
                        "Box weight (lb)", value=box_packing_optimizer.DEFAULT_BOX_SPEC["box_weight_lb"], min_value=0.1, key="default_box_weight"
                    )
                    box_packing_optimizer.DEFAULT_BOX_SPEC["box_length_in"] = bc3.number_input(
                        "Box length (in)", value=box_packing_optimizer.DEFAULT_BOX_SPEC["box_length_in"], min_value=1.0, key="default_box_length"
                    )
                    box_packing_optimizer.DEFAULT_BOX_SPEC["box_width_in"] = bc4.number_input(
                        "Box width (in)", value=box_packing_optimizer.DEFAULT_BOX_SPEC["box_width_in"], min_value=1.0, key="default_box_width"
                    )

                n_default_spec = (packing_plan_df["Box Spec Source"] == "Default (no spec on manifest)").sum()
                if n_default_spec > 0:
                    st.warning(
                        f"⚠️ {n_default_spec} row(s) used the default box spec because the manifest didn't "
                        f"include box info for that SKU. Edit the default above, or add box specs to your "
                        f"manifest for exact figures."
                    )

                st.dataframe(packing_plan_df, width='stretch', hide_index=True)

                box_total = packing_plan_df["Total Boxes"].sum()
                st.caption(f"**{int(box_total)} total boxes** across all regions for this shipment.")

                st.markdown("###### Per-Region Files — Box Counts Pre-Filled, Exact Manifest Format")
                box_export_cols = st.columns(3)
                for i, region in enumerate(REGIONS):
                    box_export = box_packing_optimizer.build_region_box_manifest_export(packing_plan_df, region)
                    if len(box_export) == 0:
                        box_export_cols[i].caption(f"No units allocated to {region}.")
                        continue
                    box_xlsx_bytes = manifest_intelligence.build_exact_format_manifest_xlsx(box_export)
                    box_export_cols[i].download_button(
                        f"⬇️ {region} — {int(box_export['Number of boxes'].sum())} boxes",
                        data=box_xlsx_bytes,
                        file_name=f"manifest_{region.lower()}_with_boxes_{datetime.now().strftime('%Y%m%d')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key=f"manifest_box_dl_{region}",
                    )
            else:
                st.info("No SKUs on this manifest matched your sales history — nothing to recommend yet.")

            if len(unmatched_df) > 0:
                with st.expander(f"Unmatched SKUs ({len(unmatched_df)}) — no sales history found"):
                    st.dataframe(unmatched_df, width='stretch', hide_index=True)
    elif st.session_state.active_manifest is not None:
        st.info(
            f"Currently using a previously uploaded manifest: "
            f"**{st.session_state.active_manifest['filename']}**. Upload a new file above to replace it, "
            f"or use the **Clear manifest** button at the top of any page to remove it."
        )
    else:
        st.info("Upload a manifest file to generate a shipment plan.")

# ============================ EXPORT PAGE ===================================
if active_page == "export":
    if manifest_mode:
        st.success(
            f"📦 **Manifest Mode active** — every export below covers only the "
            f"{len(manifest_scope_skus)} SKU(s) on **{st.session_state.active_manifest.get('filename', 'your manifest')}**."
        )

    st.markdown("#### Export Recommendations")
    st.caption("Generates fresh from current data and assumptions — not a static snapshot of an old run.")

    e1, e2 = st.columns(2)
    bulk_units = e1.number_input("Units to recommend per SKU (applies to all SKUs in export)",
                                  min_value=1, value=1000, step=50)
    bulk_size_tier = e2.selectbox("Size tier (applies to all SKUs)", list(st.session_state.fee_schedule.keys()),
                                   index=1, key="bulk_tier")
    bulk_weight = st.number_input("Avg unit weight (lb, applies to all SKUs)", min_value=0.01, value=1.1, step=0.1,
                                   key="bulk_weight")

    export_demand_df = demand_profile_scoped
    export_label = (
        f"Manifest: {st.session_state.active_manifest['filename']}" if manifest_mode
        else "All SKUs (full catalog)"
    )

    sku_units_map = {sku: bulk_units for sku in export_demand_df["SKU"].tolist()}
    rec_table = decision_engine.build_recommendation_table(
        export_demand_df, freight_rates, bulk_size_tier, bulk_weight, sku_units_map,
        fee_schedule=st.session_state.fee_schedule
    )

    st.dataframe(rec_table, width='stretch', hide_index=True)

    rec_cost_bar = go.Figure(data=[go.Bar(
        x=rec_table["SKU"], y=rec_table["Recommended Total Cost ($)"],
        marker_color=ORANGE, text=[f"${v:,.0f}" for v in rec_table["Recommended Total Cost ($)"]], textposition="outside",
    )])
    rec_cost_bar.update_layout(
        height=320, plot_bgcolor="white", paper_bgcolor="white",
        font=dict(family="Georgia, serif", color=CHARCOAL),
        yaxis_title="Recommended cost ($)", margin=dict(t=20, b=20, l=50, r=20), showlegend=False,
    )
    st.plotly_chart(rec_cost_bar, width='stretch')

    wb = excel_export.build_workbook(
        rec_table, export_demand_df, freight_rates, st.session_state.fee_schedule,
        window_days, as_of_date, sku_filter_label=export_label
    )
    excel_bytes = excel_export.workbook_to_bytes(wb)

    st.download_button(
        "⬇️ Download Excel Workbook (Summary + Full Detail)",
        data=excel_bytes,
        file_name=f"shipment_intelligence_{datetime.now().strftime('%Y%m%d')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    st.markdown("---")
    st.markdown("#### Export Manifest-Based Shipment Plan")
    st.caption("Full workbook for the active manifest — summary, region rows, box packing, and unmatched SKUs.")

    if st.session_state.active_manifest is None or len(st.session_state.active_manifest.get("summary_df", pd.DataFrame())) == 0:
        st.warning("Go to the Shipment Plan page and upload a manifest first.")
    else:
        plan = st.session_state.active_manifest
        manifest_wb = excel_export.build_workbook(
            plan["summary_df"].rename(columns={"Merchant SKU": "SKU"}), demand_profile, freight_rates,
            st.session_state.fee_schedule, window_days, as_of_date,
            sku_filter_label=f"Manifest: {plan['filename']}"
        )
        from openpyxl.styles import Font
        ws_regions = manifest_wb.create_sheet("Region Rows (Long Format)")
        ws_regions.cell(row=1, column=1, value="One row per SKU per region — ready to filter into per-region uploads").font = Font(bold=True)
        for j, col in enumerate(plan["region_rows_df"].columns, start=1):
            ws_regions.cell(row=3, column=j, value=col).font = Font(bold=True)
        for i, (_, r) in enumerate(plan["region_rows_df"].iterrows()):
            for j, col in enumerate(plan["region_rows_df"].columns, start=1):
                ws_regions.cell(row=4 + i, column=j, value=r[col])

        if len(plan.get("packing_plan_df", pd.DataFrame())) > 0:
            ws_packing = manifest_wb.create_sheet("Box Packing Plan")
            ws_packing.cell(row=1, column=1, value="Box-level breakdown per SKU per region").font = Font(bold=True)
            for j, col in enumerate(plan["packing_plan_df"].columns, start=1):
                ws_packing.cell(row=3, column=j, value=col).font = Font(bold=True)
            for i, (_, r) in enumerate(plan["packing_plan_df"].iterrows()):
                for j, col in enumerate(plan["packing_plan_df"].columns, start=1):
                    ws_packing.cell(row=4 + i, column=j, value=r[col])

        if len(plan["unmatched_df"]) > 0:
            ws_unmatched = manifest_wb.create_sheet("Unmatched SKUs")
            ws_unmatched.cell(row=1, column=1, value="SKUs on the manifest with no sales history on record").font = Font(bold=True)
            for j, col in enumerate(plan["unmatched_df"].columns, start=1):
                ws_unmatched.cell(row=3, column=j, value=col).font = Font(bold=True)
            for i, (_, r) in enumerate(plan["unmatched_df"].iterrows()):
                for j, col in enumerate(plan["unmatched_df"].columns, start=1):
                    ws_unmatched.cell(row=4 + i, column=j, value=r[col])

        manifest_excel_bytes = excel_export.workbook_to_bytes(manifest_wb)
        st.download_button(
            "⬇️ Download Manifest Shipment Plan Workbook",
            data=manifest_excel_bytes,
            file_name=f"manifest_shipment_plan_{datetime.now().strftime('%Y%m%d')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    st.markdown("---")
    st.markdown("#### Export PowerPoint Snapshot")
    st.caption("2-slide snapshot for the SKU currently selected in the Recommendation tab.")

    if "_last_recommendation" not in st.session_state:
        st.warning("Go to the Recommendation tab and select a SKU first.")
    else:
        snap = st.session_state["_last_recommendation"]
        pptx_bytes = pptx_export.build_snapshot(
            sku_label=snap["sku"], region_split=snap["region_split"], demand_pct=snap["demand_pct"],
            recommended_cost=snap["recommended_cost"], cheapest_region=snap["cheapest_region"],
            cheapest_cost=snap["cheapest_cost"], cost_delta=snap["cost_delta"],
            coverage_gap_pct=snap["coverage_gap_pct"], rationale=snap["rationale"],
            window_days=window_days, as_of_date=as_of_date,
        )
        st.download_button(
            f"⬇️ Download PowerPoint Snapshot — {snap['sku']}",
            data=pptx_bytes,
            file_name=f"shipment_snapshot_{snap['sku']}_{datetime.now().strftime('%Y%m%d')}.pptx",
            mime="application/vnd.openxmlformats-officedocument.presentationml.presentation",
        )

# ============================ DATA & SETTINGS PAGE ===========================
if active_page == "setup":
    theme.render_eyebrow(st, "Configuration")
    st.markdown("## Data & Settings")
    st.caption("Upload your files and set assumptions here — everything else in the left rail uses this data.")

    st.markdown("#### 1. Data Source")
    use_sample_input = st.checkbox(
        "Use sample data (Tuffo rain suits)", key="use_sample"
    )

    if not use_sample_input:
        if data_persistence.has_persisted_data():
            meta_display = data_meta["uploaded_at_display"] if data_meta else "unknown date"
            age_str = f"{data_age_days} day(s) ago" if data_age_days is not None else ""
            age_color = "🟢" if (data_age_days is None or data_age_days <= 7) else "🟠"
            st.success(
                f"{age_color} **Data loaded — last uploaded {meta_display}** ({age_str}). "
                f"No need to re-upload — this data is used automatically every session until you replace it."
            )
            if data_age_days is not None and data_age_days > 7:
                st.warning(
                    "This data is more than a week old. If it's past your usual Monday refresh, "
                    "consider replacing it below so recommendations reflect current sell-through."
                )

            with st.expander("Replace this week's data"):
                st.caption("Uploading new files here overwrites the previous week's data completely.")
                c1, c2, c3 = st.columns(3)
                new_sales = c1.file_uploader("Sales History (.xlsx/.csv)", type=["xlsx", "csv"], key="_replace_sales")
                new_inventory = c2.file_uploader("Inventory by Region (.xlsx/.csv)", type=["xlsx", "csv"], key="_replace_inv")
                new_invoices = c3.file_uploader("Shipment Invoices (.xlsx/.csv)", type=["xlsx", "csv"], key="_replace_invoices")
                if st.button("Save and Replace Data", type="primary"):
                    if all([new_sales, new_inventory, new_invoices]):
                        data_persistence.save_uploaded_files(new_sales, new_inventory, new_invoices)
                        st.success("Data replaced. Reloading...")
                        st.rerun()
                    else:
                        st.error("Upload all 3 files before replacing — partial replacement isn't supported, "
                                  "since the demand and cost engines need all three to stay consistent with each other.")
        else:
            st.info(
                "No data uploaded yet. Upload your Sales History, Inventory, and Invoices once — "
                "they'll be saved automatically and reused every session until you upload new ones "
                "(e.g. your weekly Monday refresh)."
            )
            c1, c2, c3 = st.columns(3)
            sales_file = c1.file_uploader("Sales History (.xlsx/.csv)", type=["xlsx", "csv"], key="_initial_sales")
            inventory_file = c2.file_uploader("Inventory by Region (.xlsx/.csv)", type=["xlsx", "csv"], key="_initial_inv")
            invoices_file = c3.file_uploader("Shipment Invoices (.xlsx/.csv)", type=["xlsx", "csv"], key="_initial_invoices")
            if st.button("Save Data", type="primary"):
                if all([sales_file, inventory_file, invoices_file]):
                    data_persistence.save_uploaded_files(sales_file, inventory_file, invoices_file)
                    st.success("Data saved. Reloading...")
                    st.rerun()
                else:
                    st.error("Upload all 3 files before saving.")
    else:
        st.success("Using built-in sample data (Tuffo rain suits, 14 months of synthetic history).")

    st.markdown("---")
    st.markdown("#### 2. Demand Window")
    st.slider(
        "Trailing days of sales history", 30, 180, step=15,
        key="window_days"
    )

    st.markdown("---")
    st.markdown("#### 3. Demand Basis")
    st.toggle(
        "Use sell-through-corrected demand (recommended)",
        key="use_corrected",
        help="OFF uses raw historical unit volume, which can create a feedback loop: regions that were "
             "overstocked look like 'high demand' even if they're just slow-moving. ON corrects for this "
             "using sell-through velocity (units sold relative to units available)."
    )

    st.markdown("---")
    st.markdown("#### 4. Placement Fee Schedule ($/unit)")
    st.caption("Editable — verify against current Seller Central rates. Amazon revises these periodically.")
    for tier in st.session_state.fee_schedule:
        st.markdown(f"**{tier}**")
        c1, c2, c3 = st.columns(3)
        st.session_state.fee_schedule[tier]["1_location"] = c1.number_input(
            "1 location", value=float(st.session_state.fee_schedule[tier]["1_location"]),
            key=f"{tier}_1loc_setup", step=0.05, format="%.2f"
        )
        st.session_state.fee_schedule[tier]["2-4_locations"] = c2.number_input(
            "2-4 locations", value=float(st.session_state.fee_schedule[tier]["2-4_locations"]),
            key=f"{tier}_24loc_setup", step=0.05, format="%.2f"
        )
        st.session_state.fee_schedule[tier]["5+_locations"] = c3.number_input(
            "5+ locations", value=float(st.session_state.fee_schedule[tier]["5+_locations"]),
            key=f"{tier}_5loc_setup", step=0.05, format="%.2f"
        )

    st.markdown("---")
    st.info(
        "Changes here apply immediately across every page — Dashboard, Recommendation, Shipment Plan, "
        "and Export all recompute live from whatever is set on this page."
    )

