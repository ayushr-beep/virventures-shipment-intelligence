"""
Excel Export.

Produces a workbook with:
  - Summary tab: clean recommendation table, ready to action
  - Demand Profile tab: full detail (raw vs corrected demand pct, velocity)
  - Cost Model tab: derived freight rates with source invoice counts
  - Assumptions tab: placement fee schedule used, as of date, window days
    (so the export is self-documenting if it lands in someone else's inbox)

Follows xlsx skill conventions: professional font, blue=inputs,
black=formulas/values, currency/percent formatting, documented sources.
"""
import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

FONT_NAME = "Calibri"
HEADER_FILL = PatternFill("solid", start_color="33373D", end_color="33373D")
HEADER_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(name=FONT_NAME, bold=True, color="33373D", size=14)
SUBTITLE_FONT = Font(name=FONT_NAME, italic=True, color="5B6168", size=10)
BODY_FONT = Font(name=FONT_NAME, size=10.5)
ACCENT_FILL = PatternFill("solid", start_color="FBEEE3", end_color="FBEEE3")
ORANGE_FONT = Font(name=FONT_NAME, bold=True, color="B85A18", size=10.5)
THIN_BORDER = Border(bottom=Side(style="thin", color="D8DADD"))


def _style_header_row(ws, row_idx, n_cols):
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=row_idx, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _autosize_columns(ws, df, start_col=1):
    for i, col in enumerate(df.columns):
        col_letter = get_column_letter(start_col + i)
        max_len = max(len(str(col)), df[col].astype(str).map(len).max() if len(df) else 0)
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 45)


def _write_title_block(ws, title, subtitle, row=1):
    ws.cell(row=row, column=1, value=title).font = TITLE_FONT
    ws.cell(row=row + 1, column=1, value=subtitle).font = SUBTITLE_FONT


def _write_dataframe(ws, df, start_row, currency_cols=None, pct_cols=None):
    currency_cols = currency_cols or []
    pct_cols = pct_cols or []

    for j, col_name in enumerate(df.columns, start=1):
        ws.cell(row=start_row, column=j, value=col_name)
    _style_header_row(ws, start_row, len(df.columns))

    for i, (_, record) in enumerate(df.iterrows()):
        r = start_row + 1 + i
        for j, col_name in enumerate(df.columns, start=1):
            val = record[col_name]
            cell = ws.cell(row=r, column=j, value=val)
            cell.font = BODY_FONT
            cell.border = THIN_BORDER
            if col_name in currency_cols:
                cell.number_format = '$#,##0.00;($#,##0.00);"-"'
            elif col_name in pct_cols:
                cell.number_format = '0.0"%"'

    return start_row + 1 + len(df)


def build_workbook(recommendation_df, demand_profile_df, freight_rates_df,
                    fee_schedule: dict, window_days, as_of_date, sku_filter_label="All SKUs"):
    wb = Workbook()

    # --- Summary tab ---
    ws = wb.active
    ws.title = "Recommendation Summary"
    _write_title_block(
        ws, "Master Shipment Intelligence — Recommendation Summary",
        f"Generated {datetime.now().strftime('%Y-%m-%d %H:%M')}  |  Scope: {sku_filter_label}  |  "
        f"Demand window: trailing {window_days} days (as of {as_of_date})"
    )
    currency_cols = [c for c in recommendation_df.columns if "($)" in c]
    pct_cols = [c for c in recommendation_df.columns if "%" in c]
    end_row = _write_dataframe(ws, recommendation_df, start_row=4, currency_cols=currency_cols, pct_cols=pct_cols)
    _autosize_columns(ws, recommendation_df)
    ws.freeze_panes = "A5"

    note_row = end_row + 2
    ws.cell(row=note_row, column=1,
            value="Note: This is a recommendation, not an executed shipment. Review the Rationale "
                  "column and current cost assumptions (Assumptions tab) before creating the shipment "
                  "in Seller Central.").font = SUBTITLE_FONT

    # --- Demand Profile detail tab ---
    ws2 = wb.create_sheet("Demand Profile (Detail)")
    _write_title_block(
        ws2, "Regional Demand Profile — Full Detail",
        "Raw % = share of historical unit volume.  Corrected % = sell-through-rate weighted "
        "(adjusts for regions that were over/under-stocked, so they aren't over/under-credited)."
    )
    display_cols = [c for c in demand_profile_df.columns]
    pct_cols2 = [c for c in display_cols if "pct" in c]
    _write_dataframe(ws2, demand_profile_df, start_row=4, pct_cols=pct_cols2)
    _autosize_columns(ws2, demand_profile_df)
    ws2.freeze_panes = "A5"

    # --- Cost Model tab ---
    ws3 = wb.create_sheet("Cost Model")
    _write_title_block(
        ws3, "Freight Cost Model — Derived From Invoice History",
        "rate_per_lb and rate_per_unit are BACK-CALCULATED from your historical shipment invoices "
        "(total_cost / total_weight or total_units), not estimated placeholders."
    )
    currency_cols3 = ["total_cost", "rate_per_lb", "rate_per_unit"]
    _write_dataframe(ws3, freight_rates_df, start_row=4, currency_cols=currency_cols3)
    _autosize_columns(ws3, freight_rates_df)

    # --- Assumptions tab ---
    ws4 = wb.create_sheet("Assumptions")
    ws4.cell(row=1, column=1, value="Assumptions Used in This Export").font = TITLE_FONT
    ws4.cell(row=2, column=1, value=f"Generated {datetime.now().strftime('%Y-%m-%d %H:%M')}").font = SUBTITLE_FONT

    r = 4
    ws4.cell(row=r, column=1, value="Demand window (trailing days)").font = Font(name=FONT_NAME, bold=True)
    ws4.cell(row=r, column=2, value=window_days).font = Font(name=FONT_NAME, color="0000FF")
    r += 1
    ws4.cell(row=r, column=1, value="As-of date").font = Font(name=FONT_NAME, bold=True)
    ws4.cell(row=r, column=2, value=str(as_of_date)).font = Font(name=FONT_NAME, color="0000FF")
    r += 2

    ws4.cell(row=r, column=1, value="Amazon Placement Fee Schedule ($/unit) — EDITABLE, verify against "
                                     "current Seller Central rates before relying on this for decisions"
             ).font = Font(name=FONT_NAME, bold=True, italic=True)
    r += 1
    headers = ["Size Tier", "1 Location", "2-4 Locations", "5+ Locations"]
    for j, h in enumerate(headers, start=1):
        ws4.cell(row=r, column=j, value=h)
    _style_header_row(ws4, r, len(headers))
    r += 1
    for tier, fees in fee_schedule.items():
        ws4.cell(row=r, column=1, value=tier).font = BODY_FONT
        ws4.cell(row=r, column=2, value=fees["1_location"]).number_format = '$#,##0.00'
        ws4.cell(row=r, column=3, value=fees["2-4_locations"]).number_format = '$#,##0.00'
        ws4.cell(row=r, column=4, value=fees["5+_locations"]).number_format = '$#,##0.00'
        for c in (2, 3, 4):
            ws4.cell(row=r, column=c).font = Font(name=FONT_NAME, color="0000FF")
        r += 1

    for col, width in zip("ABCD", [22, 14, 16, 14]):
        ws4.column_dimensions[col].width = width

    return wb


def workbook_to_bytes(wb):
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
