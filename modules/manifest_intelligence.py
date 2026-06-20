"""
Manifest Intelligence.

Reads Amazon's "Send to Amazon" manifest upload template (the real file
format used to declare shipment contents in Seller Central), extracts
SKU + Quantity per line, and runs each SKU through the existing demand
+ cost engine to produce a full shipment-level regional split plan.

This does NOT replace the manifest -- it reads it, and ALSO produces
ready-to-use region-specific manifest files in the same Send-to-Amazon
column format so the user can upload each region's portion directly.
"""
import pandas as pd
import numpy as np
import io
from . import decision_engine

REGIONS = ["East", "Central", "West"]

MANIFEST_TEMPLATE_SHEET_CANDIDATES = [
    "Create workflow – template", "Create workflow - template", "Create workflow template"
]
EXPECTED_MANIFEST_COLS = [
    "Merchant SKU", "Quantity", "Expiration date (MM/DD/YYYY)",
    "Manufacturing lot code", "Units per box", "Number of boxes",
    "Box length (in)", "Box width (in)", "Box height (in)", "Box weight (lb)",
]


def _find_header_row(raw_df, max_scan=15):
    """Scans the first N rows for the row that contains 'Merchant SKU' as a
    cell value -- that's the true header row in Amazon's template, since
    the sheet has merged/grouped label rows above it that pandas can't
    parse as column names directly."""
    for i in range(min(max_scan, len(raw_df))):
        row_values = raw_df.iloc[i].astype(str).str.strip().tolist()
        if "Merchant SKU" in row_values:
            return i
    return None


def _locate_manifest_sheet_and_header(xl):
    """
    Shared detection logic: finds the right sheet and header row in a
    Send-to-Amazon manifest workbook. Returns (sheet_name, header_row_idx).
    Raises ValueError with a human-readable message if not found -- used
    by both parse_manifest (SKU+Quantity only) and parse_manifest_full
    (all columns, including box specs) so the two stay consistent.
    """
    sheet_name = None
    for candidate in MANIFEST_TEMPLATE_SHEET_CANDIDATES:
        if candidate in xl.sheet_names:
            sheet_name = candidate
            break
    if sheet_name is None:
        for name in xl.sheet_names:
            raw = xl.parse(name, header=None)
            if _find_header_row(raw) is not None:
                sheet_name = name
                break
    if sheet_name is None:
        raise ValueError(
            "Couldn't find a 'Merchant SKU' column in any sheet of this file. "
            "This doesn't look like a Send to Amazon manifest template -- "
            "please double check the file."
        )

    raw = xl.parse(sheet_name, header=None)
    header_row_idx = _find_header_row(raw)
    if header_row_idx is None:
        raise ValueError(
            f"Found the sheet '{sheet_name}' but couldn't locate the 'Merchant SKU' "
            "header row within the first 15 rows. The template format may have changed."
        )
    return sheet_name, header_row_idx


def parse_manifest_full(file_obj):
    """
    Like parse_manifest, but returns ALL columns from the manifest --
    including box-pack columns (Units per box, Number of boxes, Box
    length/width/height, Box weight) when present -- instead of
    collapsing to just Merchant SKU + Quantity. Used by the box packing
    optimizer, which needs these columns; parse_manifest stays narrow
    for callers that only care about SKU + total quantity.
    """
    xl = pd.ExcelFile(file_obj)
    sheet_name, header_row_idx = _locate_manifest_sheet_and_header(xl)

    df = xl.parse(sheet_name, header=header_row_idx)
    df.columns = [str(c).strip() for c in df.columns]

    if "Merchant SKU" not in df.columns:
        raise ValueError(f"Found a header row, but 'Merchant SKU' column is missing. Columns found: {list(df.columns)}")

    df = df[df["Merchant SKU"].notna()].copy()
    df["Merchant SKU"] = df["Merchant SKU"].astype(str).str.strip()
    return df


def parse_manifest(file_obj):
    """
    Reads an uploaded Send-to-Amazon manifest (.xlsx) and returns a clean
    DataFrame with columns: Merchant SKU, Quantity (+ any optional columns
    present). Auto-detects the header row and the correct sheet, since the
    template wraps the real header in instructional rows above it.

    Raises ValueError with a human-readable message if the expected
    structure (a 'Merchant SKU' header) can't be found anywhere.
    """
    xl = pd.ExcelFile(file_obj)
    sheet_name, header_row_idx = _locate_manifest_sheet_and_header(xl)

    df = xl.parse(sheet_name, header=header_row_idx)
    df.columns = [str(c).strip() for c in df.columns]

    if "Merchant SKU" not in df.columns or "Quantity" not in df.columns:
        raise ValueError(
            "Found a header row, but 'Merchant SKU' and/or 'Quantity' columns are missing. "
            f"Columns found: {list(df.columns)}"
        )

    df = df[df["Merchant SKU"].notna()].copy()
    df["Merchant SKU"] = df["Merchant SKU"].astype(str).str.strip()
    df["Quantity"] = pd.to_numeric(df["Quantity"], errors="coerce")
    df = df[df["Quantity"].notna() & (df["Quantity"] > 0)]
    df["Quantity"] = df["Quantity"].astype(int)

    # Collapse duplicate SKU rows (manifest allows up to 4 packing lines per
    # MSKU -- e.g. different expiration dates). For demand/cost splitting we
    # care about TOTAL quantity per SKU; box-level detail isn't computed in
    # this version, so duplicates are summed.
    collapsed = df.groupby("Merchant SKU", as_index=False)["Quantity"].sum()
    return collapsed


def match_skus_to_demand_profile(manifest_df, demand_profile_df, sku_col="SKU"):
    """
    Matches manifest SKUs against the known demand profile. Returns
    (matched_df, unmatched_skus) -- unmatched SKUs have no sales history on
    record, so the engine has no demand signal for them and they must be
    flagged rather than silently defaulted.
    """
    known_skus = set(demand_profile_df[sku_col].astype(str))
    manifest_df = manifest_df.copy()
    manifest_df["_matched"] = manifest_df["Merchant SKU"].isin(known_skus)

    matched = manifest_df[manifest_df["_matched"]].drop(columns="_matched")
    unmatched = manifest_df[~manifest_df["_matched"]].drop(columns="_matched")
    return matched, unmatched


def build_shipment_plan(manifest_df, demand_profile_df, freight_rates, size_tier,
                         avg_unit_weight, fee_schedule=None, sku_col="SKU"):
    """
    Runs every matched SKU in the manifest through recommend_split and
    returns two DataFrames:

      summary_df: one row per SKU, columns = East/Central/West units (wide format)
      region_rows_df: one row per SKU PER REGION with nonzero units (long format,
                       ready to filter into per-region Send-to-Amazon files)

    Also returns unmatched_skus (DataFrame) for SKUs with no demand history,
    and a totals dict for the overall shipment.
    """
    matched, unmatched = match_skus_to_demand_profile(manifest_df, demand_profile_df, sku_col)

    summary_rows = []
    region_rows = []

    for _, row in matched.iterrows():
        sku = row["Merchant SKU"]
        qty = int(row["Quantity"])
        profile_row = demand_profile_df[demand_profile_df[sku_col] == sku].iloc[0]
        demand_pct = {r: float(profile_row[f"{r}_demand_pct_corrected"]) for r in REGIONS}

        result = decision_engine.recommend_split(
            qty, demand_pct, freight_rates, size_tier, avg_unit_weight, fee_schedule=fee_schedule
        )
        units = result["demand_optimal"]["units"]

        summary_row = {"Merchant SKU": sku, "Total Quantity": qty}
        for r in REGIONS:
            summary_row[f"{r} Units"] = units[r]
            summary_row[f"{r} Demand %"] = round(demand_pct.get(r, 0) * 100, 1)
        summary_row["Recommended Cost ($)"] = result["demand_optimal"]["cost"]["total_cost"]
        summary_row["Cheapest-Only Cost ($)"] = result["cheapest"]["cost"]["total_cost"]
        summary_row["Cost Delta ($)"] = result["cost_delta_vs_cheapest"]
        summary_row["Rationale"] = result["rationale"]
        summary_rows.append(summary_row)

        for r in REGIONS:
            if units[r] > 0:
                region_rows.append({
                    "Region": r, "Merchant SKU": sku, "Quantity": units[r],
                    "Demand %": round(demand_pct.get(r, 0) * 100, 1),
                })

    summary_df = pd.DataFrame(summary_rows)
    region_rows_df = pd.DataFrame(region_rows, columns=["Region", "Merchant SKU", "Quantity", "Demand %"])

    totals = {
        "total_skus": len(matched),
        "unmatched_skus": len(unmatched),
        "total_units": int(matched["Quantity"].sum()) if len(matched) else 0,
        "total_recommended_cost": round(summary_df["Recommended Cost ($)"].sum(), 2) if len(summary_df) else 0.0,
        "total_cheapest_cost": round(summary_df["Cheapest-Only Cost ($)"].sum(), 2) if len(summary_df) else 0.0,
    }
    totals["total_cost_delta"] = round(totals["total_recommended_cost"] - totals["total_cheapest_cost"], 2)

    return summary_df, region_rows_df, unmatched, totals


def build_region_manifest_export(region_rows_df, region):
    """
    Builds a DataFrame in the SAME column layout as Amazon's Send-to-Amazon
    template (Merchant SKU, Quantity, + blank optional columns) filtered to
    one region -- ready to paste into that region's shipment template.
    """
    region_df = region_rows_df[region_rows_df["Region"] == region][["Merchant SKU", "Quantity"]].copy()
    for col in EXPECTED_MANIFEST_COLS:
        if col not in region_df.columns:
            region_df[col] = np.nan
    return region_df[EXPECTED_MANIFEST_COLS]


# Amazon's REAL header row text, verbatim -- including trailing spaces on
# some columns, confirmed against an actual downloaded manifest template.
# This matters specifically for the exact-format export below: reproducing
# these strings byte-for-byte (not the cleaned/stripped versions used
# elsewhere in this module for robust READING) is what makes the output
# file genuinely re-uploadable to Seller Central without Amazon's own
# validation flagging a header mismatch.
REAL_MANIFEST_HEADERS = [
    "Merchant SKU", "Quantity", "Expiration date (MM/DD/YYYY)",
    "Manufacturing lot code ", "Units per box ", "Number of boxes",
    "Box length (in)", "Box width (in)", "Box height (in)", "Box weight (lb)",
]
REAL_MANIFEST_INSTRUCTION_ROW0 = "Please review the Example tab before you complete this sheet"
REAL_MANIFEST_OPTIONAL_ROW4 = {2: "Optional", 4: "Optional: Use only for case-packed SKUs"}
REAL_MANIFEST_SHEET_NAME = "Create workflow – template"


def build_exact_format_manifest_xlsx(export_df):
    """
    Builds an actual .xlsx file that reproduces Amazon's real Send-to-Amazon
    template structure EXACTLY: same sheet name, same instructional row 0,
    same blank spacer rows, same "Optional" markers on row 4, same header
    row (with Amazon's real trailing-space quirks) on row 5, data starting
    row 6 -- not just a CSV with matching column names.

    export_df: a DataFrame with columns matching EXPECTED_MANIFEST_COLS
    (the cleaned/stripped names) -- this function maps them onto Amazon's
    real header strings on the way out.

    Returns raw .xlsx bytes, ready for st.download_button.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = REAL_MANIFEST_SHEET_NAME

    # Row 1: instructional banner
    ws.cell(row=1, column=1, value=REAL_MANIFEST_INSTRUCTION_ROW0)
    ws.cell(row=1, column=1).font = Font(italic=True, color="666666")

    # Rows 2-4: blank spacers (rows 2,3,4 in 1-indexed = original rows 1,2,3)
    # Row 5 (1-indexed) = original row 4: "Optional" markers
    for col_idx, text in REAL_MANIFEST_OPTIONAL_ROW4.items():
        ws.cell(row=5, column=col_idx + 1, value=text)
        ws.cell(row=5, column=col_idx + 1).font = Font(italic=True, color="888888")

    # Row 6 (1-indexed) = original row 5: the real header row
    header_row_excel = 6
    for j, header_text in enumerate(REAL_MANIFEST_HEADERS, start=1):
        cell = ws.cell(row=header_row_excel, column=j, value=header_text)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", start_color="E8E8E8", end_color="E8E8E8")

    # Data rows starting row 7 (1-indexed) = original row 6 onward
    # Map export_df's cleaned column names -> position in REAL_MANIFEST_HEADERS
    clean_to_real = {clean.strip(): real for clean, real in zip(EXPECTED_MANIFEST_COLS, REAL_MANIFEST_HEADERS)}
    ordered_clean_cols = EXPECTED_MANIFEST_COLS  # export_df is already in this order

    for i, (_, row) in enumerate(export_df.iterrows()):
        excel_row = header_row_excel + 1 + i
        for j, clean_col in enumerate(ordered_clean_cols, start=1):
            val = row.get(clean_col, None)
            if pd.isna(val):
                val = None
            ws.cell(row=excel_row, column=j, value=val)

    # Reasonable column widths so it's immediately readable on open
    widths = [20, 10, 22, 20, 13, 14, 14, 13, 14, 14]
    for j, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=j).column_letter].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
